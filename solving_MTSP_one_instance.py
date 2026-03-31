import argparse
import os
import math
import random
import numpy as np
import openpyxl
import traceback
import time
from tqdm import tqdm
from clustering_k_means_plusplus_sklearn import k_means_plusplus_clustering_sklearn, cluster_tuning_for_one_node_cluster
from load_data_from_tsplib import loadDataFormTSPLibFile
from neural_end2end_solver import solve_mtsp_by_end2end_model
from tools import HyperparametersConfig, TSP_tour_distance



def solve_one_MTSP_with_different_K(data_file,model,standard_cluster_size,k_list,alpha,log_tag,doLevel2Clustering=True):
    if None == data_file:
        print("data_file is None. return None")
        return None
    
        
    timestamp=time.strftime("%Y%m%d%H%M%S", time.localtime())
    k_list=[k_list[0]] + k_list
    data_file_name=data_file.split('/')[-1][0:-4]
    
    mtsp_data_file_list=[data_file]

    # load config parameters
    config=HyperparametersConfig("config.ini")

    alpha_list = [alpha]

    ### record each mstp, under each k, the time consumed.
    each_mstp_each_k_time_consumed_list=[]

    for mtsp in tqdm(mtsp_data_file_list):
        different_k_result=[]

        ## load data here, load only once for each mtsp.
        _,data=loadDataFormTSPLibFile(mtsp)
        #### time list
        #### save the avg time consumed under each k. avg time = Total_time/len(alpha_list)
        fixed_k_time_consumed_list=[]
        for k in tqdm(k_list):
            fixed_k_scaling_factor_result=[]

            ### save the time consumed under each alpha.
            with_alpha_time_consumed_list=[]
            for scaling_factor in alpha_list:
                ### record the starting time, in million second.
                with_alpha_time_start=int(round(time.time()*1000))
                ### check at the end. if the two is the same, it means exception occured.
                with_alpha_time_end=with_alpha_time_start

                results=[]

                ### load data only once. 
                data=np.array(data)
                random_seed=int(time.time())
                clusters=k_means_plusplus_clustering_sklearn(data, k, max_iterations=config.max_clustering_iter, n_jobs=config.n_jobs, random_seed=random_seed)
                ### Refine clusters by resolving singletons
                clusters=cluster_tuning_for_one_node_cluster(clusters,True)


                cluster_data_list=[]  
                for c in clusters:
                    cluster_data_list.append(np.copy(c["points"]))  

                try:
                    if config.pre_normalization:
                        do_normalization=False
                    else:
                        do_normalization=True
                    results=solve_mtsp_by_end2end_model(cluster_data_list,model,standard_cluster_size,scaling_factor,config.cluster_algo,normalize=do_normalization,max_iterations=config.max_clustering_iter,n_jobs=config.n_jobs,doLevel2Clustering=doLevel2Clustering)
                except:
                    print("[{}] !!! Error occured when testing scaling_factor [{}].\n".format(mtsp,scaling_factor))
                    # print exception message
                    
                    traceback.print_exc()
                    ## -1 means error occured
                    result_tour_length=-1
                    ### error occured, end time set to be -1
                    with_alpha_time_end=-1

                if len(results) != len(clusters):
                    print("[{}] !!! Error occured when testing scaling_factor [{}].\n".format(mtsp,scaling_factor))
                    ## -1 means error occured
                    result_tour_length=-1
                    with_alpha_time_end=-1

                total_points=0
                for r in results:
                    total_points+=len(r)   
                    ## config init parameters
                if total_points != len(data):
                    print("[{}] !!! Error occured when testing scaling_factor [{}].\n".format(mtsp,scaling_factor))
                    with_alpha_time_end=-1

                total_distance=0
                for r in results:
                    total_distance+=TSP_tour_distance(r)

                ## get the end time, using million seconds
                if with_alpha_time_end != -1:
                    with_alpha_time_end=int(round(time.time()*1000))
                    with_alpha_time_consumed_list.append(with_alpha_time_end-with_alpha_time_start)
                
                
                print("----------cluster sizes: k={}---------------".format(k))
                for r in results:
                    print("cluster size:{}".format(len(r)))
                print("----------end cluster sizes---------------")
                
                fixed_k_scaling_factor_result.append(total_distance)
            
            # save to list
            different_k_result.append(fixed_k_scaling_factor_result)

            ## save time consumed under k.
            ## the time consumed is the mean of with_alpha_time_consumed_list
            if len(with_alpha_time_consumed_list) == 0:
                fixed_k_time_consumed_list.append(-1)
            else:
                fixed_k_time_consumed_list.append(sum(with_alpha_time_consumed_list)/len(with_alpha_time_consumed_list))
        
        each_mstp_each_k_time_consumed_list.append(fixed_k_time_consumed_list)
    
    different_k_result=np.array(different_k_result)
    
    # save to excel file
    excel=openpyxl.Workbook()

    for mtspIndex,mtsp in enumerate(mtsp_data_file_list):
        sheet=excel.create_sheet(mtsp.split('/')[-1][0:-4], index=0)
        for row in range(1,len(different_k_result)):
            sheet.cell(row=1,column=row+1,value="k={}".format(k_list[row]))
            for col in range(len(different_k_result[row])):
                sheet.cell(row=2+col,column=1,value="alpha={}".format(alpha_list[col]))
                sheet.cell(row=2+col,column=row+1,value=different_k_result[row][col])
            
            ## add a cell, print time consumed.
            sheet.cell(row=2+len(alpha_list),column=1,value="Time")
            sheet.cell(row=2+len(alpha_list),column=1+row,value=each_mstp_each_k_time_consumed_list[mtspIndex][row])
    
    # save the excel file    
    if log_tag is not None:
        excel_file_name="log/"+data_file_name+"_"+log_tag+"_result_"+timestamp+".xlsx"
    else:
        excel_file_name="log/"+data_file_name+"_result_"+timestamp+".xlsx"

    excel.save(excel_file_name)



if __name__ == "__main__":

    config=HyperparametersConfig("config.ini")

    model=config.model
    standard_cluster_size=config.standard_cluster_size
    alpha=config.scaling_factor
    

    parser = argparse.ArgumentParser(
        description="Run MTSP experiment with user-provided data_file and k_list."
    )
    parser.add_argument(
        "--data_file", "-d", required=True, type=str,
        help="Path to the TSP/MTSP data file, e.g., data/eil76.tsp"
    )
    parser.add_argument(
        "--k_list", "-k", required=True, nargs='+',type=int, 
        help="List of cluster sizes, e.g. -k 2 3 4 5"
    )

    args = parser.parse_args()

    data_file = args.data_file
    if not os.path.exists(data_file):
        raise FileNotFoundError(f"Data file not found: {data_file}")

    k_list = args.k_list

    print("data_file:", data_file)
    print("k_list:", k_list)

    
    pure_data_file_name=data_file.split("/")[-1].split(".")[0]
    
    k_list_str="_".join([str(k) for k in k_list])
    log_tag="{}_TSP{}Model_k_{}_multi_deposit_level2Clustering".format(pure_data_file_name, standard_cluster_size, k_list_str)

    solve_one_MTSP_with_different_K(data_file,model,standard_cluster_size,k_list,alpha,log_tag,doLevel2Clustering=True)
    