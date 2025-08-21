
import os
import numpy as np
import openpyxl
from tqdm import tqdm
from clustering_k_means_plusplus_sklearn import k_means_plusplus_clustering_sklearn, cluster_tuning_for_one_node_cluster
from load_data_from_tsplib import loadDataFormTSPLibFile
from neural_end2end_solver import solve_mtsp_by_end2end_model
from tools import HyperparametersConfig, TSP_tour_distance
import time


def test_TSPLIB_symmetric_EUC_2D_instance(data_file,model,standard_cluster_size,k_list,max_alpha,log_tag,log_folder=None):
    if None == data_file:
        print("data_file is None. return None")
        return None
    
    # get timestamp in format YYYYMMDDHHMMSS
    timestamp=time.strftime("%Y%m%d%H%M%S", time.localtime())
    

    ## get the pure file name, without path and suffix
    data_file_name=data_file.split('/')[-1][0:-4]
    
    mtsp_data_file_list=[data_file]

    # load config parameters
    config=HyperparametersConfig("config.ini")

    alpha_list = [max_alpha]

    each_mstp_each_k_time_consumed_list=[]

    for mtsp in tqdm(mtsp_data_file_list):
        different_k_result=[]

        ## load data here, load only once for each mtsp.
        _,data=loadDataFormTSPLibFile(mtsp)
        #### time list
        #### save the avg time consumed under each k. avg time = Total_time/len(alpha_list)
        fixed_k_time_consumed_list=[]
        for k in tqdm(k_list):
            fixed_k_scaling_factor_result=[]

            ### save the time consumed under each alpha.
            with_alpha_time_consumed_list=[]
            for scaling_factor in alpha_list:
                ### record the starting time, in million second.
                with_alpha_time_start=int(round(time.time()*1000))
                ### check at the end. if the two is the same, it means exception occured.
                with_alpha_time_end=with_alpha_time_start

                results=[]

                ### load data only once. 
                data=np.array(data)
                clusters=k_means_plusplus_clustering_sklearn(data, k, max_iterations=config.max_clustering_iter, n_jobs=config.n_jobs)
                ### tuning the clusters, make sure each cluster has at least 2 points.
                clusters=cluster_tuning_for_one_node_cluster(clusters,True)


                cluster_data_list=[]  # each element is a numpy array, 2d, [[x,y],[x,y],...]
                for c in clusters:
                    cluster_data_list.append(np.copy(c["points"]))  # copy the data, so that the original data will not be changed 

                try:
                    if config.pre_normalization:
                        do_normalization=False
                    else:
                        do_normalization=True
                    results=solve_mtsp_by_end2end_model(cluster_data_list,model,standard_cluster_size,scaling_factor,config.cluster_algo,normalize=do_normalization,max_iterations=config.max_clustering_iter,n_jobs=config.n_jobs)
                except:
                    print("[{}] !!! Error occured when testing scaling_factor [{}].\n".format(mtsp,scaling_factor))
                    # print exception message
                    import traceback
                    traceback.print_exc()
                    ## -1 means error occured
                    result_tour_length=-1
                    ### error occured, end time set to be -1
                    with_alpha_time_end=-1

                if len(results) != len(clusters):
                    print("[{}] !!! Error occured when testing scaling_factor [{}].\n".format(mtsp,scaling_factor))
                    ## -1 means error occured
                    result_tour_length=-1
                    with_alpha_time_end=-1

                total_points=0
                for r in results:
                    total_points+=len(r)   
                    ## config init parameters
                if total_points != len(data):
                    print("[{}] !!! Error occured when testing scaling_factor [{}].\n".format(mtsp,scaling_factor))
                    with_alpha_time_end=-1

                total_distance=0
                for r in results:
                    total_distance+=TSP_tour_distance(r)

                ## get the end time, using million seconds
                if with_alpha_time_end != -1:
                    with_alpha_time_end=int(round(time.time()*1000))
                    with_alpha_time_consumed_list.append(with_alpha_time_end-with_alpha_time_start)
                
                
                print("----------cluster sizes: k={}---------------".format(k))
                for r in results:
                    print("cluster size:{}".format(len(r)))
                print("----------end cluster sizes---------------")

                if log_folder is None or log_folder == "":
                    ### get time_stamp, format is : yyyyMMddTHHmmss
                    time_stamp=time.strftime("%Y%m%d%H%M", time.localtime())
                    log_folder="log/tsplib95SymmetricEuc2dAll_"+time_stamp

                if not os.path.exists(log_folder):
                    os.makedirs(log_folder)

                
                # save to log file.
                if log_tag is not None:
                    cluster_cities_log_file=log_folder+"/"+data_file_name+"_"+log_tag+"_cluster_cicites_k_"+str(k)+"_"+timestamp+".txt"
                else:
                    cluster_cities_log_file=log_folder+"/"+data_file_name+"_cluster_cities_"+str(k)+"-"+timestamp+".txt"

                # save to log file.
                with open(cluster_cities_log_file,'w') as f:
                    for r in results:
                        f.write("cluster size:{}".format(len(r)))
                        f.write("\n")


                # save to list
                fixed_k_scaling_factor_result.append(total_distance)
            
            # save to list
            different_k_result.append(fixed_k_scaling_factor_result)

            ## save time consumed under k.
            if len(with_alpha_time_consumed_list) == 0:
                fixed_k_time_consumed_list.append(-1)
            else:
                fixed_k_time_consumed_list.append(sum(with_alpha_time_consumed_list)/len(with_alpha_time_consumed_list))
        
        each_mstp_each_k_time_consumed_list.append(fixed_k_time_consumed_list)

    
    # convert to numpy array
    different_k_result=np.array(different_k_result)
    
    # save to excel file
    excel=openpyxl.Workbook()

    for mtspIndex,mtsp in enumerate(mtsp_data_file_list):
        sheet=excel.create_sheet(mtsp.split('/')[-1][0:-4])
        for row in range(len(different_k_result)):
            sheet.cell(row=1,column=row+2,value="k={}".format(k_list[row]))
            for col in range(len(different_k_result[row])):
                sheet.cell(row=2+col,column=1,value="alpha={}".format(alpha_list[col]))
                sheet.cell(row=2+col,column=row+2,value=different_k_result[row][col])
            
            ## add a cell, print time consumed.
            sheet.cell(row=2+len(alpha_list),column=1,value="Time")
            sheet.cell(row=2+len(alpha_list),column=2+row,value=each_mstp_each_k_time_consumed_list[mtspIndex][row])
    
    # save the excel file
    
    if log_tag is not None:
        excel_file_name=log_folder+"/"+data_file_name+"_"+log_tag+"_result_"+timestamp+".xlsx"
    else:
        excel_file_name=log_folder+"/"+data_file_name+"_result_"+timestamp+".xlsx"

    excel.save(excel_file_name)    


if __name__ == "__main__":

    # load config parameters
    config=HyperparametersConfig("config.ini")
        
    TSP_data_file_path=config.TSP_data_file_path

    model=config.model
    model_size=config.standard_cluster_size
    alpha=config.scaling_factor

    max_small_scale=config.max_small_scale
    max_medium_scale=config.max_medium_scale

    file_list_small_scale=[]
    file_list_medium_scale=[]
    file_list_large_scale=[]


    # for file in os.listdir("data/tsplib95_all_symmetic_ecu_2d_tsp"):
    for file in os.listdir(TSP_data_file_path):
        ## only get the file with suffix .tsp
        if file.endswith(".tsp"):
            digitial_index=0
            for i in range(len(file)):
                if file[i].isdigit():
                    digitial_index=i
                    break
            node_number=int(file[digitial_index:-4])

            if node_number<= max_small_scale:
                # file_list_small_scale.append("data/tsplib95_all_symmetic_ecu_2d_tsp/"+file)
                file_list_small_scale.append(TSP_data_file_path+"/"+file)
            elif node_number<=max_medium_scale:
                # file_list_medium_scale.append("data/tsplib95_all_symmetic_ecu_2d_tsp/"+file)
                file_list_medium_scale.append(TSP_data_file_path+"/"+file)
            else:
                # file_list_large_scale.append("data/tsplib95_all_symmetic_ecu_2d_tsp/"+file)
                file_list_large_scale.append(TSP_data_file_path+"/"+file)


    pure_data_file_name_list_small_scale=[]
    for file in file_list_small_scale:
        pure_data_file_name_list_small_scale.append(file.split("/")[-1].split(".")[0])
    
    pure_data_file_name_list_medium_scale=[]
    for file in file_list_medium_scale:
        pure_data_file_name_list_medium_scale.append(file.split("/")[-1].split(".")[0])
    
    pure_data_file_name_list_large_scale=[]
    for file in file_list_large_scale:
        pure_data_file_name_list_large_scale.append(file.split("/")[-1].split(".")[0])


    ### log file, log success count, error tsp
    ### create a new log file, add timestamp
    time_stamp=time.strftime("%Y%m%d%H%M%S", time.localtime())

    ### create a folder under log/, the folder name is "tsplibtsplib95SymmetricEuc2dAll_"+time_stamp
    # log_folder="log/tsplib95SymmetricEuc2dAll_tsplib95_standard_nint_"+time_stamp
    log_folder=config.log_folder+"_"+time_stamp
    if not os.path.exists(log_folder):
        os.makedirs(log_folder)

    log_file_name=log_folder+"/solving_all_symmetric_Euc_2d_tsps_with_standard_nint_"+time_stamp+".log"
    ### open log file
    log_file=open(log_file_name,'w')
    ## save time
    log_file.write("\nTime:{}".format(time_stamp))
    
    ### to obtain the accurate execution time, repeat the first k value twice.

    k_list_small_scale= [int(config.k_list_small_scale[0])]  # make sure the first k is always included
    for k in config.k_list_small_scale:
        k_list_small_scale.append(int(k))
    config.k_list_small_scale=k_list_small_scale
    # similarly for medium and large scale
    k_list_medium_scale=[int(config.k_list_medium_scale[0])]
    for k in config.k_list_medium_scale:
        k_list_medium_scale.append(int(k))
    config.k_list_medium_scale=k_list_medium_scale
    
    k_list_large_scale=[int(config.k_list_large_scale[0])]
    for k in config.k_list_large_scale:
        k_list_large_scale.append(int(k))
    config.k_list_large_scale=k_list_large_scale

    ## save k_list: small, medium, large
    log_file.write("\nk_list_small_scale:{}".format(k_list_small_scale))
    log_file.write("\nk_list_medium_scale:{}".format(k_list_medium_scale))
    log_file.write("\nk_list_large_scale:{}".format(k_list_large_scale))

    log_file.write("\n\n==================Small scale================================\n")

    ## convert to string to name the log file. format is k1_k2_k3_k4
    k_list_str_small_scale="_".join([str(k) for k in k_list_small_scale])
    k_list_str_medium_scale="_".join([str(k) for k in k_list_medium_scale])
    k_list_str_large_scale="_".join([str(k) for k in k_list_large_scale])

    error_tsp_list=[]
    success_count=0
    
       
    for i in range(len(file_list_small_scale)):
        data_file=file_list_small_scale[i]
        pure_data_file_name=pure_data_file_name_list_small_scale[i]
        try:
            test_TSPLIB_symmetric_EUC_2D_instance(data_file,model,model_size,k_list_small_scale,alpha,log_tag="{}_TSP{}Model_k_{}".format(pure_data_file_name,model_size,k_list_str_small_scale),log_folder=log_folder)
            print("************************************")
            print("Small scale Success:{}".format(data_file))
            print("************************************")
            ## save to log file
            log_file.write("\n{}".format(pure_data_file_name))
            success_count+=1
        except:
            error_tsp_list.append(data_file)
            print("Error occured when processing small scale TSP:{}".format(data_file))
            import traceback
            traceback.print_exc()
    
    print("Small scal Success count:{}".format(success_count))
    print("Small scale Error count:{}".format(len(error_tsp_list)))
    print("Small scale Error list:{}".format(error_tsp_list))

    ## save to log file
    log_file.write("\nSmall scal Success count:{}".format(success_count))
    log_file.write("\nSmall scale Error count:{}".format(len(error_tsp_list)))
    log_file.write("\nSmall scale Error list:{}".format(error_tsp_list))

    log_file.write("\n\n=================Medium Scale=================================\n")
    

    
    print("======================Medium scale========================")
    error_tsp_list=[]
    success_count=0

    ## for medium scale TSPs
    for i in range(len(file_list_medium_scale)):
        data_file=file_list_medium_scale[i]
        pure_data_file_name=pure_data_file_name_list_medium_scale[i]
        try:
            test_TSPLIB_symmetric_EUC_2D_instance(data_file,model,model_size,k_list_medium_scale,alpha,log_tag="{}_TSP{}Model_k_{}".format(pure_data_file_name,model_size,k_list_str_medium_scale),log_folder=log_folder)
            print("************************************")
            print("Medium scale Success:{}".format(data_file))
            print("************************************")
            ## save to log file
            log_file.write("\n{}".format(pure_data_file_name))
            success_count+=1
        except:
            error_tsp_list.append(data_file)
            print("Error occured when processing medium scale TSP:{}".format(data_file))
            import traceback
            traceback.print_exc()
    
    print("Medium scal Success count:{}".format(success_count))
    print("Medium scale Error count:{}".format(len(error_tsp_list)))
    print("Medium scale Error list:{}".format(error_tsp_list))

    ## save to log file
    log_file.write("\nMedium scal Success count:{}".format(success_count))
    log_file.write("\nMedium scale Error count:{}".format(len(error_tsp_list)))
    log_file.write("\nMedium scale Error list:{}".format(error_tsp_list))

    log_file.write("\n\n=================Large scale=================================\n")

    

    print("=====================large scale========================")
    error_tsp_list=[]
    success_count=0
    ## for large scale TSPs
    for i in range(len(file_list_large_scale)):
        data_file=file_list_large_scale[i]
        pure_data_file_name=pure_data_file_name_list_large_scale[i]
        try:
            test_TSPLIB_symmetric_EUC_2D_instance(data_file,model,model_size,k_list_large_scale,alpha,log_tag="{}_TSP{}Model_k_{}".format(pure_data_file_name,model_size, k_list_str_large_scale),log_folder=log_folder)
            print("************************************")
            print("Large scale Success:{}".format(data_file))
            print("************************************")
            ## save to log file
            log_file.write("\n{}".format(pure_data_file_name))
            success_count+=1
        except:
            error_tsp_list.append(data_file)
            print("Error occured when processing large scale TSP:{}".format(data_file))
            import traceback
            traceback.print_exc()
    
    print("Large scal Success count:{}".format(success_count))
    print("Large scale Error count:{}".format(len(error_tsp_list)))
    print("Large scale Error list:{}".format(error_tsp_list))

     ## save to log file
    log_file.write("\nLarge scal Success count:{}".format(success_count))
    log_file.write("\nLarge scale Error count:{}".format(len(error_tsp_list)))
    log_file.write("\nLarge scale Error list:{}".format(error_tsp_list))

    ## close log file
    log_file.close()

        
    